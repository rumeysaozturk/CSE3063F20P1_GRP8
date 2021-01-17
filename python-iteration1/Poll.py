import pandas as pd
import csv
import Student
import Answer
import Question
import xlsxwriter
import matplotlib.pyplot as plt
import xlwt
import xlrd
import os.path
from os import path
import openpyxl as xl;

class Poll ():
    def __init__(self, answerLength, answers):
        self.questionsInPoll = [] #1 poll için soru texti, kimlerin hangi şıkkı kaç kez seçtiği, seçilen tüm şıklar
        self.questionText = []
        self.answers = answers # soruların cevap anahtarı- doğru cevapları
        self.answerLength = answerLength

    def writePollInformation(self, students, readPath, studentsLength, Polls):

        iter = 0
        length_Polls = len(Polls)
        while iter < length_Polls:
            row = 0
            col = 0
            newPath = readPath.split(".csv")
            if iter == 0:
                picname=newPath[0]
            readPath = str(iter+1) + "Poll_Results_" + newPath[0]
            workbook = xlsxwriter.Workbook(readPath + '.xlsx')
            worksheet = workbook.add_worksheet()
            # Iterate over the data and write it out row by row.

            worksheet.write(row, col, "Student Number")
            col += 1
            worksheet.write(row, col, "Student Name")
            col += 1
            for questionNumber in range(1, Polls[iter].answerLength + 1):
                worksheet.write(row, col, "Q" + str(questionNumber))
                col += 1

            worksheet.write(row, col, "Total Questions")
            col += 1
            worksheet.write(row, col, "Correct Answers")
            col += 1
            worksheet.write(row, col, "Success Rate")
            col += 1
            worksheet.write(row, col, "SuccessPercentage")
            col += 1

            i = 0;
            while i < len(students):
                row += 1
                col = 0
                worksheet.write(row, col, str(students[i].studentid))
                col += 1
                worksheet.write(row, col, students[i].firstName + " " + students[i].lastName)
                col += 1
                j = 0
                while j < len(students[i].question[iter]):
                    worksheet.write(row, col, students[i].question[iter][j])
                    col += 1
                    j += 1


                worksheet.write(row, col, Polls[iter].answerLength)
                col += 1
                worksheet.write(row, col, students[i].totalpointforThispoll[iter])
                col += 1
                worksheet.write(row, col, str(students[i].totalpointforThispoll[iter]) + " of " + str(Polls[iter].answerLength))
                col += 1
                worksheet.write(row, col, str((students[i].totalpointforThispoll[iter] / Polls[iter].answerLength) * 100))
                col += 1
                i += 1

            a = 0
            lengthOfPreviousAnswer = 0
            previousPlace = studentsLength + 4 + 15 * a + lengthOfPreviousAnswer
            worksheet.write('A' + str(previousPlace), "----------------QUESTIONS---------------- ")
            previousPlace += 2


            while a < Polls[iter].answerLength:
                x_bar = []
                y_bar = []
                lengthOfAnswer = len(Polls[iter].questionsInPoll[a].answers)
                i = 0
                trueAnswerIndex = 0
                while i < lengthOfAnswer:
                    x_bar.append(str(i + 1))
                    y_bar.append(Polls[iter].questionsInPoll[a].countOfStudent[i])
                    if Polls[iter].questionsInPoll[a].answers[i] == Polls[iter].answers[a].answer:
                        trueAnswerIndex = i
                    i += 1
                plt.bar(x_bar, y_bar, color='b').patches[trueAnswerIndex].set_color('r')
                nameOfPng = str(1 + a) + "Pic_"+str(iter+1)+"Poll_"+picname+".png"
                plt.savefig(nameOfPng, dpi=150)
                worksheet.write('A' + str(previousPlace), str(a + 1) + "Q) " + Polls[iter].questionsInPoll[a].question)
                previousPlace += 1
                worksheet.insert_image('B' + str(previousPlace), nameOfPng, {'x_scale': 0.5, 'y_scale': 0.5})

                i = 0
                previousPlace += 12
                while i < lengthOfAnswer:
                    previousPlace += 1
                    worksheet.write('A' + str(previousPlace),
                                    str(i + 1) + "- " + Polls[iter].questionsInPoll[a].answers[i] + " (" + str(Polls[iter].questionsInPoll[a].countOfStudent[i]) + " students select)")

                    i += 1
                previousPlace += 3
                plt.close()

                a += 1

            workbook.close()
            iter += 1

    def writeinGlobal(self,students,readPath,Polls):
       if (path.exists("Global.xlsx")):
           '''''
           loc = ("Global.xlsx")
           wb = xlrd.open_workbook(loc)
           sheet = wb.sheet_by_index(0) '''

           filename = "Global.xlsx"
           wb1 = xl.load_workbook(filename)
           ws1 = wb1.worksheets[0]

           # opening the destination excel file
           filename1 = "Global.xlsx"
           wb2 = xl.load_workbook(filename1)
           ws2 = wb2.active

           colunm = ws1.cell(1,200).value



           y=1;
           t=1;


           # add_sheet is used to create sheet.

           while y<=len(students):
               while t<=colunm:
                   c=ws1.cell(y,t).value
                   ws2.cell(y,t).value=c
                   t=t+1
               y=y+1

           colunm+=1




           i = 0;
           k=2;
           iter=0;
           length_Polls = len(Polls)
           while iter < length_Polls:
                newPath = readPath.split(".csv")
                readPath1 = str(iter + 1) +" of "+str(len(Polls)) + newPath[0]
                ws2.cell(1, colunm).value=readPath1
                while i < len(students):
                   ws2.cell(k,colunm).value=(str(Polls[iter].answerLength)+" ; "+str((students[i].totalpointforThispoll[iter] / Polls[iter].answerLength) * 100))
                   i=i+1
                   k=k+1
                i=0
                k=2
                colunm+=1
                iter += 1

           ws2.cell(1,200).value=colunm-1
           wb2.save("Global.xlsx")
       else:
          colunm=3
          wb = xlsxwriter.Workbook('Global.xlsx')
          sheet1 = wb.add_worksheet()

        # add_sheet is used to create sheet.

          sheet1.write(0, 0, "Öğrenci No")
          sheet1.write(0, 1, "Öğrenci Adı")
          sheet1.write(0, 2, "Öğrenci Soyadı")
          d = 0
          i = 1
          while d < len(students):
                sheet1.write(i, 0, str(students[d].studentid))
                sheet1.write(i, 1, students[d].firstName)
                sheet1.write(i, 2, students[d].lastName)
                d = d + 1
                i = i + 1


          i = 0;
          k=1
          iter = 0
          length_Polls = len(Polls)
          while iter < length_Polls:
            newPath = readPath.split(".csv")
            readPath1 = str(iter + 1) + " of " + str(len(Polls)) + newPath[0]
            sheet1.write(0, colunm, readPath1)
            while i < len(students):
                   sheet1.write(k, colunm, str(Polls[iter].answerLength) + " ; " + str((students[i].totalpointforThispoll[iter] / Polls[iter].answerLength) * 100))
                   i = i + 1
                   k=k+1
            k=1
            i=0
            colunm += 1
            iter += 1



          sheet1.write(0, 199, colunm)

          wb.close()
